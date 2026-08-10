import openpyxl, csv, sys, os

SRC = r"D:/工作文档zr/培训/数字化赋能资源库/专家资源/专家资源库_69位信息待修复.xlsx"
OUTDIR = os.path.join(os.path.dirname(__file__))

wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
print("SHEETS:", wb.sheetnames)
for ws in wb.worksheets:
    print("\n===== SHEET:", ws.title, "dims:", ws.max_row, "x", ws.max_column, "=====")
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        print("(empty)")
        continue
    print("HEADER:", [str(h) for h in header])
    # sample 3 data rows
    for i, r in enumerate(rows):
        if i >= 3:
            break
        print(f"ROW{i+1}:", [str(c)[:40] if c is not None else "" for c in r])
    # export full to csv
    csv_path = os.path.join(OUTDIR, "experts_export_" + ws.title + ".csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([str(h) for h in header])
        # re-iterate: read_only worksheet can't restart; reopen for full dump
    # full dump (reopen non-readonly to allow full iteration safely)
    wb2 = openpyxl.load_workbook(SRC, data_only=True)
    ws2 = wb2[ws.title]
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        for ri, row in enumerate(ws2.iter_rows(values_only=True), 1):
            w.writerow(["" if c is None else c for c in row])
    print("EXPORTED ->", csv_path, "rows:", ws2.max_row)
